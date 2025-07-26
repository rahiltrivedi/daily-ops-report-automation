import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os
import smtplib
from email.message import EmailMessage

# ---------- CONFIG ----------
INPUT_FILE = r"C:/Users/Asus/Desktop/Job/Projects/Daily Operation Summary/disbursements_2025-07-26.xlsx"
TODAY = datetime.today().strftime("%Y-%m-%d")
REPORT_FILENAME = f"daily_report_{TODAY}.xlsx"

EMAIL_SENDER = "popeye138@yahoo.com"
EMAIL_PASSWORD = "addyourapppasword"  # ⚠️ In real projects, store in environment variables or config files
EMAIL_RECIPIENT = "popeye2105@yahoo.com"
SMTP_SERVER = "smtp.mail.yahoo.com"
SMTP_PORT = 587

# ---------- STEP 1: Load and Process Data ----------
df = pd.read_excel(INPUT_FILE)
df['Date'] = pd.to_datetime(df['Date'])

today_df = df[df['Date'].dt.date == datetime.today().date()]

# Summary Metrics
total_disbursed = today_df[today_df['Status'] == 'Disbursed']['Amount'].sum()
total_rejected = today_df[today_df['Status'] == 'Rejected']['Amount'].sum()
rejected_count = today_df[today_df['Status'] == 'Rejected'].shape[0]

# Fix column names and avoid 'Reason' KeyError
rejection_reasons = today_df[today_df['Status'] == 'Rejected']['Rejected Reason'].value_counts().reset_index()
rejection_reasons.columns = ['Rejected Reason', 'Count']

# ---------- STEP 2: Generate Excel Report ----------
wb = Workbook()
ws = wb.active
ws.title = "Daily Report"

# Title
ws.append([f"Daily Disbursement Report - {TODAY}"])
ws.append([])

# Summary Table
ws.append(["Summary Metric", "Value"])
ws.append(["Total Disbursed", total_disbursed])
ws.append(["Total Rejected", total_rejected])
ws.append(["Rejected Count", rejected_count])
ws.append([])

# Rejection Reasons Table
ws.append(["Rejection Reason", "Count"])
for _, row in rejection_reasons.iterrows():
    ws.append([row['Rejected Reason'], row['Count']])
ws.append([])

# Detailed Records Table
header = ["Disbursement ID", "Customer Name", "Amount", "Status", "Rejected Reason"]
ws.append(header)

for _, row in today_df.iterrows():
    ws.append([
        row["Disbursement ID"],
        row["Customer Name"],
        row["Amount"],
        row["Status"],
        row["Rejected Reason"] if row["Status"] == "Rejected" else ""
    ])

# Style the header
for cell in ws[ws.max_row - len(today_df) - 1]:
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor="DDDDDD")

# Save the Excel report
wb.save(REPORT_FILENAME)

# ---------- STEP 3: Send Email ----------
try:
    msg = EmailMessage()
    msg["Subject"] = f"Daily Ops Report - {TODAY}"
    msg["From"] = EMAIL_SENDER
    msg["To"] = EMAIL_RECIPIENT

    msg.set_content(f"""Hello,

Please find attached the disbursement summary for {TODAY}.

Summary:
- Total Disbursed: ₹{total_disbursed:,}
- Total Rejected: ₹{total_rejected:,}
- Rejected Count: {rejected_count}

Regards,
Daily Reporting Bot
""")

    # Attach the Excel file
    with open(REPORT_FILENAME, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=REPORT_FILENAME
        )

    # Send the email
    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_SENDER, EMAIL_PASSWORD)
        smtp.send_message(msg)

    print("✅ Report created and email sent successfully.")

except Exception as e:
    print(f"❌ Failed to send email: {e}")
